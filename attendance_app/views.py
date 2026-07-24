from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.utils import timezone
from .models import Course, LectureSession, Attendance
import qrcode
import base64
from io import BytesIO
import csv
from openpyxl import Workbook
from django.urls import reverse
from django.contrib import messages
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from django.template.loader import get_template
from xhtml2pdf import pisa






def is_lecturer(user):
    return user.is_authenticated and user.is_staff

@login_required
@user_passes_test(is_lecturer)
def lecturer_dashboard(request):
    # Only show courses assigned to this lecturer
    courses = Course.objects.filter(lecturer=request.user)
    
    # Last 20 sessions by this lecturer
    sessions = LectureSession.objects.filter(
        course__lecturer=request.user
    ).select_related('course').order_by('-date')[:20]
    
    return render(request, 'attendance/lecturer_dashboard.html', {
        'courses': courses,
        'sessions': sessions
    })


@login_required
@user_passes_test(is_lecturer)
def start_session(request):
     if request.method != 'POST':
        return redirect('attendance_app:lecturer_dashboard')
     course_id = request.POST.get('course_id')
     course = get_object_or_404(Course, id=course_id, lecturer=request.user)
    # close any active for this session first
     LectureSession.objects.filter(lecturer=request.user, is_active=True).update(is_active=False)

     session = LectureSession.objects.create(course=course, lecturer=request.user, is_active=True)
     return redirect('attendance_app:session_qr', session_id=session.id) 



@login_required
@user_passes_test(is_lecturer)
def session_qr(request, session_id):
    session = get_object_or_404(LectureSession, id=session_id, lecturer=request.user)
    return render(request, 'attendance/session_qr.html', {'session':session,})




@login_required
@user_passes_test(is_lecturer)
def get_qr_image(request, session_id):
    session = get_object_or_404(LectureSession, id=session_id, lecturer=request.user)

    # Rotate token every time this is called
    session.rotate_token()

    qr_url = request.build_absolute_uri(
        reverse('attendance_app:mark_attendance', args=[session.qr_token])
    )
    qr = qrcode.make(qr_url)
    buffer = BytesIO()
    qr.save(buffer, format='PNG')
    qr_b64 = base64.b64encode(buffer.getvalue()).decode()

    return JsonResponse({
        'qr_b64':qr_b64 ,
        'token_age_sec': (timezone.now() - session.token_created_at).seconds
    })





@login_required
@user_passes_test(is_lecturer)
def end_session(request, session_id):
    session =  get_object_or_404(LectureSession, id=session_id, lecturer=request.user)
    session.is_active = False
    session.save()
    return redirect('attendance_app:lecturer_dashboard')



@login_required
@user_passes_test(is_lecturer)
def Delete_session(request, session_id):
    session =  get_object_or_404(LectureSession, id=session_id, lecturer=request.user)
    session.delete()
    return redirect('attendance_app:lecturer_dashboard')



@login_required
def scan_page(request):
    return render(request, 'Attendance/scan.html')



@login_required
def mark_attendance(request, token):
    session = get_object_or_404(LectureSession, qr_token=token)
    
    # 1. Check if session is active
    if not session.is_active:
        return render(request, 'attendance/attend_result.html', {'status': 'ended', 'session': session})
    
    # 2. Check token expiry - 2 min window
    if not session.is_token_valid():
        return render(request, 'attendance/attend_result.html', {'status': 'expired', 'session': session})
    
    # 3. Check if user is student, not lecturer
    if request.user.is_staff:
        messages.info(request, "Lecturers can't mark attendance")
        return render(request, 'attendance/attend_result.html', {'status': 'forbidden', 'session': session})
        #return HttpResponseForbidden("Lecturers can't mark attendance")
    
    # 4. Check if already marked - prevents double scan
    if Attendance.objects.filter(student=request.user, session=session).exists():
        return render(request, 'attendance/attend_result.html', {'status': 'already', 'session': session})
    
    # 5. All good - create record
    Attendance.objects.create(student=request.user, session=session)
    return render(request, 'attendance/attend_result.html', {'status': 'success', 'session': session})





@login_required
@user_passes_test(is_lecturer)
def attendance_list(request, session_id):
    session = get_object_or_404(LectureSession, id=session_id, lecturer=request.user)
    if request.GET.get('json'):
        count = Attendance.objects.filter(session=session).count()
        return JsonResponse({'count':count})
    records = Attendance.objects.filter(session=session).select_related('student')
    return render(request, 'attendance/attendance_list.html', {'records': records})







@login_required
@user_passes_test(is_lecturer)
def session_detail(request, session_id):
    session = get_object_or_404(LectureSession, id=session_id, course__lecturer=request.user)
    attendances = session.attendance_set.select_related('student', ).order_by('timestamp')
    return render(request, 'attendance/session_details.html', {
        'session': session,
        'attendances': attendances
    })

@login_required
@user_passes_test(is_lecturer)
def export_csv(request, session_id):
    session = get_object_or_404(LectureSession, id=session_id, course__lecturer=request.user)
    attendances = session.attendance_set.select_related('student', 'student__profile').order_by('timestamp')
    
    response = HttpResponse(content_type='text/csv')
    filename = f"{session.course.code}_{session.start_time.date()}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    writer.writerow(['S/N', 'Student ID', 'Full Name', 'Username', 'Time'])
    
    for i, att in enumerate(attendances, 1):
        writer.writerow([
            i,
            att.student.profile.student_id if hasattr(att.student, 'profile') else '',
            att.student.get_full_name(),
            att.student.username,
            att.timestamp.strftime('%Y-%m-%d %H:%M:%S')
        ])
    return response


@login_required
@user_passes_test(is_lecturer)
def export_excel(request, session_id):
    session = get_object_or_404(LectureSession, id=session_id, course__lecturer=request.user)
    attendances = session.attendance_set.select_related('student').order_by('timestamp')

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # 1. Add title + session info at top
    ws['A1'] = f"{session.course.code} - {session.course.name}"
    ws['A2'] = f"Date: {session.date}"
    ws['A3'] = f"Total Present: {attendances.count()}"
    ws.merge_cells('A1:D1')
    ws.merge_cells('A2:D2')
    ws.merge_cells('A3:D3')

    # Style title
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # 2. Header row
    headers = ['S/N', 'Matric Num', 'Full Name', 'Date & Time Marked']
    ws.append(headers)
    header_row = ws[5]

    # Style headers
    fill = PatternFill(start_color="1a4d8f", end_color="1a4d8f", fill_type="solid")
    for cell in header_row:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')

    # 3. Data rows
    for i, att in enumerate(attendances, 1):
        matric = att.student.student_id if att.student_id else 'N/A'
        name = att.student.get_full_name() or att.student.username
        timestamp = att.timestamp.astimezone(timezone.get_current_timezone()).strftime('%d/%m/%Y %H:%M')

        ws.append([i, matric, name, timestamp])

    # 4. Auto-adjust column widths
    column_widths = [8, 18, 30, 22]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 5. Freeze header row
    ws.freeze_panes = 'A6'

    # 6. Better filename
    date_str = session.date.strftime('%Y%m%d')
    filename = f"{session.course.code}_{date_str}_Attendance.xlsx"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response







@login_required
@user_passes_test(is_lecturer)
def export_pdf(request, session_id):
    session = get_object_or_404(LectureSession, id=session_id, course__lecturer=request.user)
    attendances = session.attendance_set.select_related('student').order_by('timestamp')

    # Convert UTC to Nigeria time WAT
    nigeria_tz = timezone.get_current_timezone()
    
    context = {
        'session': session,
        'attendances': attendances,
        'total': attendances.count(),
        'generated': timezone.now().astimezone(nigeria_tz),
        'nigeria_tz': nigeria_tz
    }

    template = get_template('attendance/lecture_session_pdf.html')
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    date_str = session.date.strftime('%Y%m%d')
    filename = f"{session.course.code}_{date_str}_Attendance.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)
    return response